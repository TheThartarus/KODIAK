from tkinter import messagebox, filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import csv

def export_data(self):
    """
    Lee el CSV y lo exporta a un archivo Excel (.xlsx).
    """
    # Verificar si el archivo CSV existe
    if not os.path.exists(self.csv_log):
        messagebox.showerror(
            "Error",
            "No hay datos para exportar.",
        )
        return False

    # Preguntar al usuario dónde guardar el archivo
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[(
            "Libro de Excel",
            "*.xlsx",
        )],
        title="Exportar datos a Excel",
    )

    # Si el usuario cancela la selección, salir de la función
    if not file_path:
        return False

    # Intentar crear el archivo Excel
    try:
        # Crear el libro y la hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "KODIAK"

        # Definir estilos
        header_font = Font(
            name="Calibri",
            size=11,
            bold=True,
            color="FFFFFF",
        )
        header_fill = PatternFill(
            start_color="2F5597",
            end_color="2F5597",
            fill_type="solid",
        )
        center_alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # Leer el CSV e insertar datos
        with open(
            self.csv_log,
            mode='r',
            encoding='utf-8'
        ) as f:
            reader = csv.reader(f)
            for r_idx, row in enumerate(reader, 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(
                        row=r_idx,
                        column=c_idx,
                        value=value,
                    )

                    # Aplicar bordes a todas las celdas con datos
                    cell.border = thin_border
                    
                    # Estilo especial para la cabecera (Fila 1)
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = Alignment(
                            vertical="center",
                        )

        # Ajustar ancho de columnas automáticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Guardar el archivo
        wb.save(file_path)

        messagebox.showinfo(
            "KODIAK",
            f"Reporte generado con éxito en:\n{file_path}",
        )
        return True

    except Exception as e:
        messagebox.showerror(
            "Error de Exportación",
            f"No se pudo crear el Excel: {e}",
        )
        return False